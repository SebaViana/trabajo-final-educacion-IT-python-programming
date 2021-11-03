import os
import time
import funciones
import openpyxl as op
from openpyxl.styles import Alignment, alignment


combo_S = 650
combo_D = 700
combo_T = 800
flurby = 250

wb = op.Workbook()

if not os.path.exists('ejemplo.xlsx'):
	
	ws1 = wb.active



	Cel1 = ws1["A1"] = "Cliente"
	Cel1 = Alignment(horizontal='center')

	Cel2 = ws1["B1"] = "Fecha"
	Cel2 = Alignment(horizontal='center')
	Cel2 = ws1.column_dimensions['B'].width = 25

	Cel3 = ws1["C1"] = "Combo S"
	Cel3 = Alignment(horizontal='center')

	Cel4 = ws1["D1"] = "Combo D"
	Cel4 = Alignment(horizontal='center')

	Cel5 = ws1["E1"] = "Combo T"
	Cel5 = Alignment(horizontal='center')

	Cel6 = ws1["F1"] = "Flurby"
	Cel6 = Alignment(horizontal='center')

	Cel7 = ws1["G1"] = "Total"
	Cel7 = Alignment(horizontal='center')

	wb.save(filename="ejemplo.xlsx")
else:
	pass

def iniciar_menu():

	while True: 
		
		
		cliente = input("Ingrese el nombre del cliente: ")
		cliente = funciones.verificar_vacio(cliente)
		cliente = funciones.verificar_nombre(cliente)
		tiempo = int(time.strftime('%H'))
		Saludo = funciones.saludo_tiempo(tiempo, cliente.capitalize())


		cant_combo_S = input("Ingrese cantidad Combo S : ")
		cant_combo_S = funciones.verificar_vacio(cant_combo_S)
		cant_combo_S = funciones.convertir(cant_combo_S)
		
		cant_combo_D = input("Ingrese cantidad Combo D : ")
		cant_combo_D = funciones.verificar_vacio(cant_combo_D)
		cant_combo_D = funciones.convertir(cant_combo_D)


		cant_combo_T = input("Ingrese cantidad Combo T : ")
		cant_combo_T = funciones.verificar_vacio(cant_combo_T)
		cant_combo_T = funciones.convertir(cant_combo_T)


		cant_flurby = input("Ingrese cantidad Flurby : ")
		cant_flurby = funciones.verificar_vacio(cant_flurby)
		cant_flurby = funciones.convertir(cant_flurby)

		
		total = combo_S * cant_combo_S + combo_D * cant_combo_D + combo_T * cant_combo_T + flurby * cant_flurby
		

		if total == 0:
			os.system("cls")
			print("No se ordenó nada. Volviendo al menu.")
			time.sleep(1)
			os.system("cls")
			print("No se ordenó nada. Volviendo al menu..")
			time.sleep(1)
			os.system("cls")
			print("No se ordenó nada. Volviendo al menu...")
			time.sleep(1)
			os.system("cls")
			break
		else:
			print("El total es de ", total, " pesos.")

		abono = input("Ingrese con cuanto desea abonar : ")
		abono = funciones.verificar_vacio(abono)
		abono = funciones.convertir(abono)
		abono = funciones.verif_vuelto(total, abono)
		
		vuelto = abono - total
		
		funciones.total_en_caja += (abono - vuelto)
		

		print("Su vuelto es de ", vuelto, " pesos.")
		print("Desea hacer la compra?")
		opcion = input("Introduzca 's' para sí y 'n' para no. >>>")
		opcion = funciones.verificar_vacio(opcion)
		opcion = funciones.verificar_nombre(opcion)
		opcion = funciones.verificar_si_no(opcion.lower())


		if opcion == "s":
			print("Gracias por comprar en Mc Dowell’s!")
			time.sleep(3)
			os.system("cls")
			print("Volviendo al menu.")
			time.sleep(1)
			os.system("cls")
			print("Volviendo al menu..")
			time.sleep(1)
			os.system("cls")
			print("Volviendo al menu...")
			time.sleep(1)
			os.system("cls")

			my_filename = "ejemplo.xlsx"

			wb = op.load_workbook(filename=my_filename)
			ws = wb['Sheet']


			newRowLocation = ws.max_row +4


			ws.append([cliente.capitalize(), time.asctime(), cant_combo_S, cant_combo_D, cant_combo_T, cant_flurby, total])
			wb.save(filename=my_filename)
			wb.close()
			break
		else:
			print("Su orden fue cancelada.")
			time.sleep(5)
			os.system("cls")
			print("Volviendo al menu.")
			time.sleep(1)
			os.system("cls")
			print("Volviendo al menu..")
			time.sleep(1)
			os.system("cls")
			print("Volviendo al menu...")
			time.sleep(1)
			os.system("cls")
			break





  


